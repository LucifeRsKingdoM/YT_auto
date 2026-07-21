import io
import os
import tempfile
import datetime as dt

from flask import (
    Blueprint, render_template, jsonify, request, send_file, session, redirect, url_for
)
from werkzeug.utils import secure_filename

from database import get_session
from models import Video, UploadFailure, ActivityLog, YouTubeAccount
from security import login_required, current_user
from storage_backend import get_storage

bp = Blueprint("videos", __name__)


@bp.route("/videos")
@login_required
def page():
    s = get_session()
    try:
        # Fetch only the logged-in user's channels for the add modal / selection UI
        user = current_user()
        user_channels = s.query(YouTubeAccount).filter_by(owner_username=user).all()
        return render_template("videos.html", channels=user_channels)
    finally:
        s.close()


@bp.route("/api/videos")
@login_required
def list_videos():
    """Tabbed list restricted to the current user. ?tab=scheduled|uploaded|failed|all&slot=<slot_number>"""
    tab = request.args.get("tab", "all")
    slot_filter = request.args.get("slot")
    user = current_user()
    
    s = get_session()
    try:
        # STRICT ISOLATION: Only fetch videos belonging to this user
        q = s.query(Video).filter(Video.owner_username == user)
        
        if slot_filter:
            q = q.join(YouTubeAccount).filter(YouTubeAccount.slot_number == int(slot_filter))
            
        if tab == "scheduled":
            q = q.filter(Video.status == "scheduled")
        elif tab == "uploaded":
            q = q.filter(Video.status == "uploaded")
        elif tab == "failed":
            q = q.filter(Video.status == "failed")
            
        rows = q.order_by(Video.updated_at.desc()).all()
        out = []
        for v in rows:
            row = _row(v)
            if tab == "failed":
                fail = (
                    s.query(UploadFailure)
                    .filter_by(video_id=v.id)
                    .order_by(UploadFailure.timestamp.desc())
                    .first()
                )
                row["failure"] = {
                    "error_message": fail.error_message,
                    "log": fail.log,
                    "timestamp": fail.timestamp.isoformat() + "Z",
                } if fail else None
            out.append(row)
        return jsonify(out)
    finally:
        s.close()


@bp.route("/api/videos", methods=["POST"])
@login_required
def create_video():
    """Create a video record tied to the current user and a specific channel slot."""
    user = current_user()
    s = get_session()
    try:
        title = request.form.get("title", "").strip()
        if not title:
            return jsonify({"error": "Title is required."}), 400
            
        # Accept either slot_number or youtube_account_id from frontend form
        slot_number = request.form.get("slot_number")
        account_id = request.form.get("youtube_account_id")

        account = None
        if slot_number:
            account = s.query(YouTubeAccount).filter_by(
                owner_username=user, slot_number=int(slot_number)
            ).first()
        elif account_id:
            account = s.query(YouTubeAccount).filter_by(
                owner_username=user, id=int(account_id)
            ).first()
            
        if not account:
            # Corrected fallback query using session query properly
            account = s.query(YouTubeAccount).filter_by(owner_username=user).first()
            if not account:
                return jsonify({"error": "Please connect a YouTube channel slot first on the Integrations page."}), 400
            
            
        v = Video(
            title=title,
            description=request.form.get("description", ""),
            tags=request.form.get("tags", ""),
            privacy=request.form.get("privacy", "private"),
            category_id=request.form.get("category_id", "22"),
            owner_username=user,
            youtube_account_id=account.id,
        )
        s.add(v)
        s.flush()  # get v.id

        file = request.files.get("file")
        if file and file.filename:
            _store_upload(v, file, user, account.slot_number)

        s.add(ActivityLog(actor=user, action="Video added", detail=f"{title} (Slot {account.slot_number})"))
        s.commit()
        return jsonify(_row(v)), 201
    except Exception as exc:  
        s.rollback()
        return jsonify({"error": str(exc)}), 400
    finally:
        s.close()


@bp.route("/api/videos/<int:vid>", methods=["PUT"])
@login_required
def update_video(vid):
    user = current_user()
    s = get_session()
    try:
        # Ensure user can only edit their own videos
        v = s.query(Video).filter_by(id=vid, owner_username=user).first()
        if not v:
            return jsonify({"error": "Not found"}), 404
            
        data = request.get_json(force=True)
        for field in ("title", "description", "tags", "privacy", "category_id"):
            if field in data:
                setattr(v, field, data[field])
                
        if "slot_number" in data:
            account = s.query(YouTubeAccount).filter_by(
                owner_username=user, slot_number=int(data["slot_number"])
            ).first()
            if account:
                v.youtube_account_id = account.id

        s.add(ActivityLog(actor=user, action="Video edited", detail=v.title))
        s.commit()
        return jsonify(_row(v))
    except Exception as exc:  
        s.rollback()
        return jsonify({"error": str(exc)}), 400
    finally:
        s.close()


@bp.route("/api/videos/<int:vid>", methods=["DELETE"])
@login_required
def delete_video(vid):
    user = current_user()
    s = get_session()
    try:
        v = s.query(Video).filter_by(id=vid, owner_username=user).first()
        if not v:
            return jsonify({"error": "Not found"}), 404
            
        if v.file_key:
            get_storage().delete_file(v.file_key)
            
        title = v.title
        s.delete(v)
        s.add(ActivityLog(actor=user, action="Video deleted", detail=title))
        s.commit()
        return jsonify({"ok": True})
    finally:
        s.close()


@bp.route("/api/videos/<int:vid>/file", methods=["POST"])
@login_required
def upload_file(vid):
    """Attach / replace the video file for an existing record with folder segregation."""
    user = current_user()
    s = get_session()
    try:
        v = s.query(Video).filter_by(id=vid, owner_username=user).first()
        if not v:
            return jsonify({"error": "Not found"}), 404
            
        file = request.files.get("file")
        if not file or not file.filename:
            return jsonify({"error": "No file provided"}), 400
            
        slot_num = v.youtube_account.slot_number if v.youtube_account else 1
        _store_upload(v, file, user, slot_num)
        
        s.add(ActivityLog(actor=user, action="File attached", detail=v.title))
        s.commit()
        return jsonify(_row(v))
    finally:
        s.close()


# ---- Excel import / export ---------------------------------------------
@bp.route("/api/videos/export.xlsx")
@login_required
def export_xlsx():
    from openpyxl import Workbook
    user = current_user()
    s = get_session()
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "Videos"
        headers = ["id", "title", "description", "tags", "privacy",
                   "status", "youtube_video_id", "scheduled_time", "slot_number"]
        ws.append(headers)
        
        # Export only current user's videos
        rows = s.query(Video).filter_by(owner_username=user).order_by(Video.id).all()
        for v in rows:
            slot_num = v.youtube_account.slot_number if v.youtube_account else ""
            ws.append([
                v.id, v.title, v.description, v.tags, v.privacy, v.status,
                v.youtube_video_id,
                v.scheduled_time.isoformat() if v.scheduled_time else "",
                slot_num,
            ])
        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return send_file(
            buf, as_attachment=True, download_name=f"videos_{user}.xlsx",
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    finally:
        s.close()


@bp.route("/api/videos/import.xlsx", methods=["POST"])
@login_required
def import_xlsx():
    from openpyxl import load_workbook
    user = current_user()
    file = request.files.get("file")
    if not file or not file.filename:
        return jsonify({"error": "No file provided"}), 400
        
    wb = load_workbook(file, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return jsonify({"error": "Empty file"}), 400
        
    header = [str(h).strip().lower() if h else "" for h in rows[0]]
    idx = {name: header.index(name) for name in header if name}

    s = get_session()
    created = updated = 0
    try:
        for r in rows[1:]:
            def cell(name, default=""):
                i = idx.get(name)
                return (r[i] if i is not None and i < len(r) and r[i] is not None
                        else default)
            title = str(cell("title")).strip()
            if not title:
                continue
                
            slot_num = int(cell("slot_number", 1))
            account = s.query(YouTubeAccount).filter_by(
                owner_username=user, slot_number=slot_num
            ).first()
            
            v = Video(
                title=title,
                description=str(cell("description")),
                tags=str(cell("tags")),
                privacy=str(cell("privacy", "private")) or "private",
                owner_username=user,
                youtube_account_id=account.id if account else None
            )
            s.add(v)
            created += 1
            
        s.add(ActivityLog(actor=user, action="Excel import", detail=f"{created} videos added"))
        s.commit()
        return jsonify({"created": created, "updated": updated})
    except Exception as exc:  
        s.rollback()
        return jsonify({"error": str(exc)}), 400
    finally:
        s.close()


# ---- helpers ------------------------------------------------------------
def _store_upload(v, file, username, slot_number):
    filename = secure_filename(file.filename)
    # Segregate storage path cleanly by user and slot folder: user/slot_X/filename
    key = f"{username}/slot_{slot_number}/{v.id}_{filename}"
    
    tmp = tempfile.NamedTemporaryFile(delete=False)
    file.save(tmp.name)
    tmp.close()
    try:
        size = os.path.getsize(tmp.name)
        get_storage().save_file(key, tmp.name)
        v.file_key = key
        v.file_size = size
    finally:
        os.remove(tmp.name)


def _row(v):
    slot_num = v.youtube_account.slot_number if v.youtube_account else None
    channel_title = v.youtube_account.channel_name if v.youtube_account else "Unassigned"
    return {
        "id": v.id,
        "title": v.title,
        "description": v.description,
        "tags": v.tags,
        "privacy": v.privacy,
        "category_id": v.category_id,
        "status": v.status,
        "youtube_video_id": v.youtube_video_id,
        "file_key": v.file_key,
        "file_size": v.file_size,
        "slot_number": slot_num,
        "channel_name": channel_title,
        "scheduled_time": v.scheduled_time.isoformat() + "Z" if v.scheduled_time else None,
        "created_at": v.created_at.isoformat() + "Z" if v.created_at else None,
    }


@bp.route('/videos/add', methods=['GET', 'POST'])
@login_required
def add_video():
    user = current_user()
    db = get_session()
    try:
        if request.method == 'POST':
            selected_slot = request.form.get('slot_number')
            account = db.query(YouTubeAccount).filter_by(
                owner_username=user, slot_number=int(selected_slot)
            ).first()
            
            new_video = Video(
                title=request.form.get('title'),
                owner_username=user,
                youtube_account_id=account.id if account else None
            )
            db.add(new_video)
            db.commit()
            return redirect(url_for('videos.page'))
            
        # Fetch ONLY this user's connected channels for the slot picker dropdown
        user_channels = db.query(YouTubeAccount).filter_by(owner_username=user).all()
        return render_template('videos.html', channels=user_channels)
    finally:
        db.close()